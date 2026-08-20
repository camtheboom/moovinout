#!/usr/bin/env python3
"""Regenerate rentdata-vic.js from the Victorian Rental Report.

    pip install openpyxl
    python tools/build-rent-data-vic.py

Victoria's equivalent of the NSW bond data comes from the Residential Tenancies Bond
Authority, published quarterly by DFFH as the Rental Report and licensed CC BY 4.0. Like
NSW it is agreed rents rather than asking prices, and like NSW it ships bundled in the
page - no key, no proxy, no quota.

HOW THIS DIFFERS FROM THE NSW BUILD, WHICH IS MOST OF WHY IT IS A SEPARATE SCRIPT
---------------------------------------------------------------------------------
* NSW publishes per-bond rows and we compute the medians. Victoria publishes finished
  medians, already suppressed where the sample is thin (a "-"). So the two states carry
  different suppression rules and we cannot apply our own thresholds here.

* Victoria's geography is DFFH's own suburb groups, not postcodes - "Carlton-Parkville",
  "CBD-St Kilda Rd". There is no postcode to join a centroid to, so names are split and
  geocoded against the same open locality dataset the NSW build already uses, and the
  parts averaged. Two spelling conventions differ from the gazetteer often enough to be
  worth rules rather than overrides ("East Hawthorn" for Hawthorn East, "Mt Eliza" for
  Mount Eliza); the rest are listed in OVERRIDES, including one outright typo in the
  source. Anything still unresolved is reported and dropped, never guessed.

* Victoria publishes only the bedroom x dwelling-type crossing - one and two and three
  bedroom flats, two and three and four bedroom houses - plus an all-properties figure.
  It publishes no studio and no townhouse category at all, and crucially no bedroom-only
  or type-only medians. Those cannot be recovered from the crossed cells, because a
  median of medians is not a median. They are count-weighted blends instead, flagged with
  derived:true so the app can label them in the popup rather than passing them off as
  published figures. The all-properties figure is a real median and is carried as `all`.
"""

import collections
import csv
import math
import re
import statistics
import sys
import urllib.request
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is needed: pip install openpyxl")

SOURCE_PAGE = "https://www.dffh.vic.gov.au/publications/rental-report"
BASE = "https://www.dffh.vic.gov.au"
POSTCODES_CSV = "https://raw.githubusercontent.com/matthewproctor/australianpostcodes/master/australian_postcodes.csv"

ROOT = Path(__file__).resolve().parent.parent
CACHE = ROOT / "tools" / ".cache"
OUT = ROOT / "rentdata-vic.js"

# sheet name -> (bedrooms, dwelling initial). "a" is a flat/unit, matching the NSW file's
# apartment. Victoria has no studio and no townhouse sheet - that is the data, not a gap
# in this list.
SHEETS = {
    "1 bedroom flat": ("1", "a"),
    "2 bedroom flat": ("2", "a"),
    "3 bedroom flat": ("3", "a"),
    "2 bedroom house": ("2", "h"),
    "3 bedroom house": ("3", "h"),
    "4 bedroom house": ("4", "h"),
}
ALL_SHEET = "All properties"
TYPE_NAME = {"a": "apartment", "h": "house"}

DIRECTIONS = ("NORTH", "SOUTH", "EAST", "WEST")

# Names the locality gazetteer cannot be talked into on its own.
OVERRIDES = {
    "CBD": "MELBOURNE",
    "ST KILDA RD": "MELBOURNE",      # a road, not a locality; the group is the CBD end of it
    "YARRA RANGES": "LILYDALE",      # an LGA, not a locality - Lilydale is its gateway
    "WANAGARATTA": "WANGARATTA",     # misspelled in the source spreadsheet
    "NEWCOMBE": "NEWCOMB",           # same
    "BENDIGO EAST": "EAST BENDIGO",
}

# Victoria-wide, to catch a name that geocoded to the wrong side of the state. Not the
# Melbourne box - Geelong and Ballarat are legitimately in this file.
VIC_BBOX = {"minLat": -39.2, "maxLat": -33.9, "minLon": 140.9, "maxLon": 150.1}

# Two parts of one named group this far apart means one of them resolved somewhere wrong.
MAX_PART_SPREAD_KM = 40


def fetch(url, name):
    CACHE.mkdir(parents=True, exist_ok=True)
    path = CACHE / name
    if path.exists() and path.stat().st_size > 0:
        return path
    print("  downloading " + name)
    req = urllib.request.Request(url, headers={"User-Agent": "moovin-rent-data-build"})
    with urllib.request.urlopen(req, timeout=180) as r, open(path, "wb") as f:
        f.write(r.read())
    return path


def source_file():
    """Find the current moving-annual-rents-by-suburb spreadsheet.

    The slug is not stable across quarters - it has been both "rent" and "rents", and the
    quarter is spelled out - so the index page is scraped rather than a URL built.
    """
    req = urllib.request.Request(SOURCE_PAGE, headers={"User-Agent": "moovin-rent-data-build"})
    with urllib.request.urlopen(req, timeout=120) as r:
        html = r.read().decode("utf-8", "replace")
    links = re.findall(r'href="(/moving-annual-rents?-suburb-[^"]+)"', html)
    if not links:
        sys.exit("Could not find a moving-annual-rents-by-suburb link on " + SOURCE_PAGE)
    href = links[0]
    quarter = re.search(r"-(\w+)-quarter-(\d{4})", href)
    label = "%s %s" % (quarter.group(1).title()[:3], quarter.group(2)) if quarter else "unknown"
    return fetch(BASE + href, "vic-moving-annual-rent-suburb.xlsx"), label, href


def haversine_km(a, b, c, d):
    R = 6371.0
    p, q = math.radians(a), math.radians(c)
    h = (math.sin(math.radians(c - a) / 2) ** 2
         + math.cos(p) * math.cos(q) * math.sin(math.radians(d - b) / 2) ** 2)
    return 2 * R * math.asin(math.sqrt(h))


def localities():
    """VIC locality -> centroid, from the open postcode dataset the NSW build also uses."""
    path = fetch(POSTCODES_CSV, "australian_postcodes.csv")
    grouped = collections.defaultdict(list)
    with open(path, newline="", encoding="utf-8-sig") as f:
        for r in csv.DictReader(f):
            if r.get("state") != "VIC":
                continue
            try:
                lat, lon = float(r["lat"]), float(r["long"])
            except (TypeError, ValueError):
                continue
            if lat == 0:
                continue
            grouped[r["locality"].strip().upper()].append((lat, lon))
    return {k: (sum(p[0] for p in v) / len(v), sum(p[1] for p in v) / len(v))
            for k, v in grouped.items()}


def name_variants(part):
    """Spellings to try for one part of a group name, best first."""
    yield part
    if part in OVERRIDES:
        yield OVERRIDES[part]
    words = part.split()
    # DFFH writes "East Hawthorn"; the gazetteer says "Hawthorn East".
    if len(words) > 1 and words[0] in DIRECTIONS:
        yield " ".join(words[1:] + [words[0]])
    if part.startswith("MT "):
        yield "MOUNT " + part[3:]


def geocode(name, loc, problems):
    """Centroid for a DFFH suburb group, or None.

    Groups are hyphen-joined lists of localities. Each part is resolved independently and
    the results averaged; a part that lands implausibly far from the others is treated as
    a bad match and dropped rather than allowed to drag the marker across the state.
    """
    parts = [p.strip().upper() for p in name.split("-") if p.strip()]
    found, missed = [], []
    for p in parts:
        hit = next((loc[v] for v in name_variants(p) if v in loc), None)
        (found.append((p, hit)) if hit else missed.append(p))
    if not found:
        problems.append((name, "no part resolved: " + ", ".join(missed)))
        return None

    if len(found) > 1:
        centre = (sum(f[1][0] for f in found) / len(found),
                  sum(f[1][1] for f in found) / len(found))
        keep = [f for f in found
                if haversine_km(centre[0], centre[1], f[1][0], f[1][1]) <= MAX_PART_SPREAD_KM]
        if len(keep) != len(found):
            dropped = [f[0] for f in found if f not in keep]
            problems.append((name, "part(s) too far from the rest, dropped: " + ", ".join(dropped)))
            found = keep or found

    lat = sum(f[1][0] for f in found) / len(found)
    lon = sum(f[1][1] for f in found) / len(found)
    if not (VIC_BBOX["minLat"] <= lat <= VIC_BBOX["maxLat"]
            and VIC_BBOX["minLon"] <= lon <= VIC_BBOX["maxLon"]):
        problems.append((name, "resolved outside Victoria"))
        return None
    if missed:
        problems.append((name, "partly resolved, unmatched: " + ", ".join(missed)))
    return round(lat, 5), round(lon, 5)


def latest_column(sheet):
    """Index of the newest Median column, and its quarter label."""
    rows = list(sheet.iter_rows(values_only=True))
    header, sub = rows[1], rows[2]
    medians = [i for i, v in enumerate(sub) if v == "Median" and header[i]]
    if not medians:
        sys.exit("No Median column found in sheet " + sheet.title)
    i = medians[-1]
    return rows, i, str(header[i])


def number(v):
    if v is None or v == "-":
        return None
    try:
        n = float(v)
    except (TypeError, ValueError):
        return None
    return int(round(n)) if n > 0 else None


def read_sheet(sheet):
    """area name -> (median, count) for the newest quarter."""
    rows, mi, label = latest_column(sheet)
    out = {}
    for r in rows[3:]:
        name = r[1]
        if not name or str(name).strip() == "Group Total":
            continue
        median, count = number(r[mi]), number(r[mi - 1])
        if median is not None:
            out[str(name).strip()] = (median, count or 0)
    return out, label


def weighted(pairs):
    """Count-weighted blend of medians.

    This is the one place the file departs from "publish it or hide it". Victoria gives no
    bedroom-only or type-only medians and they cannot be recovered from the crossed cells,
    but dropping them would empty the map for "2 bed, any type" - the most common filter
    there is. So they are blended, and marked derived so the app can say so.
    """
    pairs = [(m, c) for m, c in pairs if m is not None and c]
    if not pairs:
        return None
    total = sum(c for _, c in pairs)
    return int(round(sum(m * c for m, c in pairs) / total))


def main():
    print("Building Victorian rent data")
    path, quarter, href = source_file()
    print("  source: %s (%s quarter)" % (href, quarter))
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    cells = {}
    for sheet_name, key in SHEETS.items():
        if sheet_name not in wb.sheetnames:
            sys.exit("Expected sheet missing: " + sheet_name)
        cells[key], _ = read_sheet(wb[sheet_name])
    overall, quarter_label = read_sheet(wb[ALL_SHEET])
    print("  %d areas in the all-properties sheet, window is the 12 months to %s"
          % (len(overall), quarter_label))

    loc = localities()
    problems = []
    areas = []
    for name in sorted(overall):
        point = geocode(name, loc, problems)
        if not point:
            continue
        lat, lon = point

        crossed = {}
        for (beds, dwelling), table in cells.items():
            hit = table.get(name)
            if hit:
                crossed[beds + dwelling] = hit[0]

        by_bed = {}
        for beds in ("1", "2", "3", "4"):
            blend = weighted([(cells[(beds, d)].get(name, (None, 0)))
                              for d in ("a", "h") if (beds, d) in cells])
            if blend is not None:
                by_bed[beds] = blend

        by_type = {}
        for dwelling in ("a", "h"):
            blend = weighted([(cells[(b, dwelling)].get(name, (None, 0)))
                              for b in ("1", "2", "3", "4") if (b, dwelling) in cells])
            if blend is not None:
                by_type[TYPE_NAME[dwelling]] = blend

        median, count = overall[name]
        if not crossed and not by_bed:
            continue
        areas.append({"n": name, "lat": lat, "lon": lon, "c": count,
                      "all": median, "bed": by_bed, "type": by_type, "x": crossed})

    print("  geocoded %d of %d areas" % (len(areas), len(overall)))
    if problems:
        print("  %d name(s) needed attention:" % len(problems))
        for name, why in problems:
            print("    %-42s %s" % (name, why))

    write(areas, quarter_label, href)


def js(o):
    if isinstance(o, str):
        return '"' + o.replace("\\", "\\\\").replace('"', '\\"') + '"'
    if isinstance(o, dict):
        return "{" + ",".join('%s:%s' % (js(str(k)), js(v)) for k, v in o.items()) + "}"
    return str(o)


def write(areas, quarter_label, href):
    lines = [
        "// Victorian median weekly rents by DFFH suburb group - GENERATED FILE, DO NOT EDIT BY HAND.",
        "// Regenerate with: python tools/build-rent-data-vic.py",
        "//",
        "// Source: the Victorian Rental Report (CC BY 4.0), built from Residential Tenancies",
        "// Bond Authority lodgements - agreed rents, not asking prices.",
        "// " + SOURCE_PAGE,
        "// Locality centroids: the open australianpostcodes dataset, averaged per locality.",
        "//",
        "// Window: the moving annual median to %s, published quarterly." % quarter_label,
        "// Suppression is DFFH's own, not ours: a cell they judged too thin is absent here.",
        "//",
        "// x: bedrooms crossed with dwelling type (\"2a\" = 2-bed flat), as published.",
        "// all: the all-properties median, as published.",
        "// bed / type: COUNT-WEIGHTED BLENDS, not published figures. Victoria publishes no",
        "// bedroom-only or type-only medians and they cannot be recovered from the crossing,",
        "// so these are derived and the app labels them as combined. Hence derived: true.",
        "//",
        "// Victoria publishes no studio and no townhouse category at all.",
        "const VIC_RENT_DATA = {",
        '  source: "Victorian Rental Report (RTBA bonds)",',
        '  to: %s,' % js(quarter_label),
        "  months: 12,",
        "  derived: true,",
        "  areas: [",
    ]
    for a in areas:
        lines.append("    {n:%s, lat:%s, lon:%s, c:%d, all:%s, bed:%s, type:%s, x:%s},"
                     % (js(a["n"]), a["lat"], a["lon"], a["c"], js(a["all"]),
                        js(a["bed"]), js(a["type"]), js(a["x"])))
    lines += ["  ],", "};", ""]
    OUT.write_text("\n".join(lines), encoding="utf-8")
    print("Wrote %s - %d areas" % (OUT.name, len(areas)))


if __name__ == "__main__":
    main()
