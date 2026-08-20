#!/usr/bin/env python3
"""Regenerate rentdata.js from NSW Fair Trading's rental bond lodgement files.

Every residential bond lodged in NSW is recorded by Fair Trading, and the raw
per-bond rows are published monthly as spreadsheets: lodgement date, postcode,
dwelling type, bedrooms, weekly rent. That is what tenants actually agreed to pay
rather than what was advertised, and it is open data - no key, no backend, no quota.

This script rolls the last MONTHS files into one set of medians per postcode, joins
a centroid so each postcode can be a map marker, and writes rentdata.js.

    pip install openpyxl
    python tools/build-rent-data.py

Run it whenever you want fresher figures; Fair Trading publishes a new month
roughly mid-month. Downloads are cached under tools/.cache so re-runs are cheap.

Two things worth knowing before changing anything here:

  * The published filenames are not consistent - "RentalBond_Lodgements_April_2023",
    "rentalbond_lodgements_september25" and "rental-bond-lodgement-data-march-2025"
    are all real. So the source page is scraped for links rather than URLs being
    built from a pattern, and the month is recovered from whatever the filename
    happens to look like.
  * Postcode centroids are not in the bond data. They come from the open
    australianpostcodes dataset, averaged over the localities sharing a postcode.
"""

import collections
import csv
import datetime as dt
import json
import re
import statistics
import sys
import urllib.request
from pathlib import Path

SOURCE_PAGE = "https://www.nsw.gov.au/housing-and-construction/rental-forms-surveys-and-data/rental-bond-data"
POSTCODES_CSV = "https://raw.githubusercontent.com/matthewproctor/australianpostcodes/master/australian_postcodes.csv"

MONTHS = 12          # rolling window; 12 evens out seasonality and thickens thin postcodes
MIN_AREA_OBS = 25    # a postcode needs this many bonds before it gets a marker at all
MIN_CELL_OBS = 8     # a single bedroom/dwelling median needs this many before it is shown
RENT_FLOOR = 50      # below this is a data-entry error, not a weekly residential rent
RENT_CEILING = 10000

# (F)lat/unit, (H)ouse, (T)errace/townhouse/semi. (O)ther is garages, car spaces and
# rented rooms, and (U) is unknown - both are dropped, so a car space can never drag
# down a bedroom median.
DWELLING = {"F": "apartment", "H": "house", "T": "townhouse"}

MONTH_NAMES = ["january", "february", "march", "april", "may", "june",
               "july", "august", "september", "october", "november", "december"]

ROOT = Path(__file__).resolve().parent.parent
CACHE = ROOT / "tools" / ".cache"


def fetch(url, name):
    """Download to the cache, or reuse what is already there."""
    CACHE.mkdir(parents=True, exist_ok=True)
    path = CACHE / name
    if path.exists() and path.stat().st_size > 0:
        return path
    print("  downloading " + name)
    req = urllib.request.Request(url, headers={"User-Agent": "moovin-rent-data-build"})
    with urllib.request.urlopen(req, timeout=180) as r, open(path, "wb") as f:
        f.write(r.read())
    return path


def monthly_files():
    """Scrape the source page and return the newest MONTHS lodgement spreadsheets."""
    req = urllib.request.Request(SOURCE_PAGE, headers={"User-Agent": "moovin-rent-data-build"})
    with urllib.request.urlopen(req, timeout=120) as r:
        html = r.read().decode("utf-8", "replace")

    found = {}
    for href in re.findall(r'href="([^"]+\.xlsx)"', html, re.I):
        name = href.rsplit("/", 1)[-1]
        key = name.lower().replace("-", "_")
        if "lodge" not in key:
            continue          # refunds files have the same columns but a different meaning
        if "year" in key:
            continue          # annual roll-ups would double-count the months around them

        month = next((i + 1 for i, m in enumerate(MONTH_NAMES) if m in key), None)
        if not month:
            continue
        # The year sits after the month name and may be 2- or 4-digit:
        # "_july_2026" but also "september25".
        tail = key.split(MONTH_NAMES[month - 1], 1)[1]
        m = re.search(r"(20\d{2})", tail) or re.search(r"(\d{2})", tail)
        if not m:
            continue
        year = int(m.group(1))
        if year < 100:
            year += 2000

        url = href if href.startswith("http") else "https://www.nsw.gov.au" + href
        found[(year, month)] = (name, url)          # later duplicates win; they are reissues

    if not found:
        sys.exit("No lodgement spreadsheets found - the source page layout has probably changed.")
    return [(ym, found[ym]) for ym in sorted(found)[-MONTHS:]]


def read_bonds(path):
    """Yield (postcode, dwelling, bedrooms, weekly rent) from one spreadsheet."""
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    # Row 1 is a title block and row 3 the header, so the data starts at row 4.
    for row in ws.iter_rows(min_row=4, values_only=True):
        cells = (list(row) + [None] * 5)[:5]
        _, postcode, dwelling, bedrooms, rent = cells
        if postcode is None:
            continue
        kind = DWELLING.get(str(dwelling).strip().upper() if dwelling else "")
        if not kind:
            continue
        try:
            weekly = float(str(rent).strip())
            beds = int(str(bedrooms).strip())
        except (TypeError, ValueError):
            continue          # both columns use "U" for unknown
        if not (RENT_FLOOR <= weekly <= RENT_CEILING):
            continue
        yield str(postcode).strip().zfill(4), kind, min(beds, 4), weekly
    wb.close()


def pick_name(localities):
    """Choose the one locality that best names a whole postcode.

    The principal suburb is usually the one the others are named after - Liverpool
    beside Liverpool South and Liverpool Westfield, Parramatta beside Parramatta
    Westfield - so the name prefixing the most siblings wins. Otherwise first
    alphabetically, which is all a postcode of unrelated villages deserves.

    This picks a real locality every time but not always the best-known one: 2042
    labels as Enmore where most people would say Newtown. Using the ABS statistical
    area name to break those ties was tried and is worse - it renames 2150 from
    Parramatta to Harris Park. Every popup shows the postcode next to the name, so a
    sibling suburb still identifies the area.
    """
    # Mail-handling and shopping-centre entries are never the answer, but they still
    # count towards the prefix score below - "Parramatta Westfield" is exactly the
    # evidence that Parramatta is the principal name here.
    candidates = [l for l in localities
                  if not re.search(r"\b(BC|DC|MC|LPO|MSC)$", l)
                  and not re.search(r"(MALL|WESTFIELD|PLAZA|PARLIAMENT HOUSE|UNIVERSITY OF)", l)]
    candidates = candidates or localities

    def prefix_score(name):
        return sum(1 for other in localities if other != name and other.startswith(name + " "))

    best = min(candidates, key=lambda n: (-prefix_score(n), n))
    return " ".join(w.capitalize() for w in best.split())


def postcode_places():
    """postcode -> (display name, lat, lon) for NSW, from the open postcode dataset."""
    path = fetch(POSTCODES_CSV, "australian_postcodes.csv")
    with open(path, encoding="utf-8-sig", newline="") as f:
        rows = [r for r in csv.DictReader(f)
                if r.get("state") == "NSW" and r.get("type") != "Post Office Boxes"]

    grouped = collections.defaultdict(list)
    for r in rows:
        try:
            lat, lon = float(r["Lat_precise"]), float(r["Long_precise"])
        except (TypeError, ValueError, KeyError):
            continue
        if lat == 0 or lon == 0:
            continue
        grouped[r["postcode"]].append((lat, lon, r["locality"].strip().upper()))

    places = {}
    for postcode, entries in grouped.items():
        # A postcode covers several localities. Averaging their points lands close to
        # the true postcode centroid - within ~500 m for the Sydney postcodes that were
        # checked against hand-verified suburb coordinates.
        lat = statistics.mean(e[0] for e in entries)
        lon = statistics.mean(e[1] for e in entries)
        name = pick_name(sorted({e[2] for e in entries}))
        places[postcode] = (name, round(lat, 5), round(lon, 5))
    return places


def main():
    print("Reading the last {} months of NSW rental bond lodgements".format(MONTHS))
    files = monthly_files()
    span = ("{}-{:02d}".format(*files[0][0]), "{}-{:02d}".format(*files[-1][0]))

    areas = collections.defaultdict(lambda: {
        "bed": collections.defaultdict(list),
        "type": collections.defaultdict(list),
        "cross": collections.defaultdict(list),
        "all": [],
    })
    total = 0
    for (year, month), (name, url) in files:
        path = fetch(url, name)
        n = 0
        for postcode, kind, beds, weekly in read_bonds(path):
            g = areas[postcode]
            g["all"].append(weekly)
            g["bed"][str(beds)].append(weekly)
            g["type"][kind].append(weekly)
            # Bedrooms and dwelling type are separate marginals, so they cannot answer
            # "2-bed apartment" between them. Keyed "2a", "3h", "0t" to stay small.
            g["cross"][str(beds) + kind[0]].append(weekly)
            n += 1
        total += n
        print("  {}-{:02d}  {:>6} bonds".format(year, month, n))
    print("  {} bonds across {} postcodes".format(total, len(areas)))

    print("Joining postcode centroids")
    places = postcode_places()

    def medians(values):
        return {k: round(statistics.median(v)) for k, v in sorted(values.items())
                if len(v) >= MIN_CELL_OBS}

    out, no_centroid = [], []
    for postcode, g in sorted(areas.items()):
        if len(g["all"]) < MIN_AREA_OBS:
            continue
        if postcode not in places:
            no_centroid.append(postcode)
            continue
        by_bed, by_type = medians(g["bed"]), medians(g["type"])
        if not by_bed and not by_type:
            continue
        name, lat, lon = places[postcode]
        out.append({"pc": postcode, "n": name, "lat": lat, "lon": lon,
                    "c": len(g["all"]), "bed": by_bed, "type": by_type,
                    "x": medians(g["cross"])})

    if no_centroid:
        print("  no centroid for {} postcodes: {}".format(
            len(no_centroid), ", ".join(no_centroid[:10])))
    print("  writing {} areas".format(len(out)))

    lines = [
        "// NSW median weekly rents by postcode - GENERATED FILE, DO NOT EDIT BY HAND.",
        "// Regenerate with: python tools/build-rent-data.py",
        "//",
        "// Source: NSW Fair Trading rental bond lodgements (CC BY 4.0) - the bonds actually",
        "// lodged for new tenancies, so these are agreed rents, not asking prices.",
        "// " + SOURCE_PAGE,
        "// Postcode centroids: the open australianpostcodes dataset, averaged per postcode.",
        "//",
        "// Window: {} to {} ({} months, {} bonds).".format(span[0], span[1], len(files), total),
        "// A postcode needs {} bonds to appear at all; a single median needs {}.".format(
            MIN_AREA_OBS, MIN_CELL_OBS),
        "// bed: 0 = studio/bedsitter, 4 = four or more. type: flat, house, terrace/townhouse.",
        "// x: the two crossed, keyed bedrooms + type initial (\"2a\" = 2-bed apartment), because",
        "// bed and type are marginals and cannot be combined after the fact.",
        "",
        "const NSW_RENT_DATA = {",
        '  source: "NSW Fair Trading rental bond lodgements",',
        '  sourceUrl: "' + SOURCE_PAGE + '",',
        '  license: "CC BY 4.0",',
        '  from: "{}", to: "{}", months: {},'.format(span[0], span[1], len(files)),
        '  generated: "{}",'.format(dt.date.today().isoformat()),
        "  observations: {},".format(total),
        "  areas: [",
    ]
    for a in out:
        lines.append(
            '    {{pc:"{pc}", n:"{n}", lat:{lat}, lon:{lon}, c:{c}, bed:{bed}, type:{type}, x:{x}}},'.format(
                pc=a["pc"], n=a["n"].replace('"', ""), lat=a["lat"], lon=a["lon"], c=a["c"],
                bed=json.dumps(a["bed"], separators=(",", ":")),
                type=json.dumps(a["type"], separators=(",", ":")),
                x=json.dumps(a["x"], separators=(",", ":"))))
    lines += ["  ],", "};", ""]

    target = ROOT / "rentdata.js"
    target.write_text("\n".join(lines), encoding="utf-8")
    print("Wrote {} ({:.0f} KB)".format(target, target.stat().st_size / 1024))


if __name__ == "__main__":
    main()
